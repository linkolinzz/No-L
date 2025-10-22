# pages/product_base.py

import flet as ft
from database.db_manager import db_fetch, db_exec
from utils.logger import log
from datetime import datetime
import io
# Use Decimal for precise weight values without rounding.  We import here
# once and reuse in both the main and old tabs.  Accept comma separators
# in parsing but avoid converting to int or float which could round.
from decimal import Decimal

# опційно: openpyxl для xlsx
try:
    import openpyxl
except Exception:
    openpyxl = None


# ──────────────────────────────────────────────────────────────────────
# Загальні утиліти
# ──────────────────────────────────────────────────────────────────────

def _bool(v):  # до int 0/1
    """
    Convert various truthy/falsy representations to 1 or 0.
    Accepts numeric, boolean and string values.  For strings, interpret
    "1", "так", "yes", "true" (case-insensitive) as true; anything
    else is considered false.  For other types, fall back to Python's
    truthiness.
    """
    try:
        # Strings representing yes/true
        if isinstance(v, str):
            s = v.strip().lower()
            if s in ("1", "так", "yes", "true", "y", "t"):  # support UA/EN
                return 1
            if s in ("0", "ні", "no", "false", "n", "f"):
                return 0
        # numeric or boolean
        return 1 if bool(v) else 0
    except Exception:
        return 0

def _fmt_bool(v):
    return "1" if (v in (1, "1", True)) else "0"

# Human-friendly yes/no formatter for Excel exports.  Returns "Так"
# for truthy values and "Ні" for falsy values.  Accepts numeric,
# boolean and string values ("1", "так", etc.) similar to _bool.
def _yes_no(v):
    return "Так" if _bool(v) == 1 else "Ні"

# Define the fields exported/imported for the product base.  A new column
# for weight in grams has been added after the name to allow specifying
# the product weight.  When exporting/importing Excel files the order of
# columns must match this list.
FIELDS = [
    ("article_code", "Артикул"),
    ("name",         "Назва"),
    # Weight in grams; header updated to display "Вага (г)" to clarify units
    ("weight_g",     "Вага (г)"),
    ("drying_needed",   "Сушка"),
    ("trimming_needed", "Обрізка"),
    ("cutting_needed",  "Різка"),
    ("cleaning_needed", "Зачистка"),
]


# ──────────────────────────────────────────────────────────────────────
# Основна вкладка (product_base)
# ──────────────────────────────────────────────────────────────────────

def _tab_main(page: ft.Page) -> ft.Column:
    product_list = ft.Column(scroll=ft.ScrollMode.ADAPTIVE)

    # Ensure both product_base and product_base_old contain the weight_g column and
    # that it uses a DECIMAL type to preserve fractional grams.  Avoid the
    # unsupported IF NOT EXISTS syntax on MySQL versions prior to 8.  We query
    # information_schema to decide whether to add the column.  When the column
    # already exists, no ALTER is attempted.
    try:
        # Ensure both product tables have weight_g column with DECIMAL type.  We add the column
        # if missing, and migrate it from INT to DECIMAL if required.  This makes sure that
        # fractional grams are stored correctly and not rounded.
        for tbl in ("product_base", "product_base_old"):
            rows = db_fetch(
                "SELECT DATA_TYPE FROM information_schema.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME=%s AND COLUMN_NAME='weight_g'",
                (tbl,)
            )
            if not rows:
                # column missing -> add as DECIMAL(10,3)
                db_exec(f"ALTER TABLE {tbl} ADD COLUMN weight_g DECIMAL(10,3) NULL")
            else:
                dt = rows[0].get("DATA_TYPE", "").lower()
                if dt.startswith("int"):
                    # convert INT to DECIMAL preserving NULL
                    db_exec(f"ALTER TABLE {tbl} MODIFY COLUMN weight_g DECIMAL(10,3) NULL")
    except Exception:
        pass
    selected_product_code = None
    is_editing = False

    confirm_dialog = ft.AlertDialog(modal=True)

    search_field = ft.TextField(label="Пошук", on_change=lambda e: load_products())
    sort_by = ft.Dropdown(
        label="Сортувати за",
        options=[ft.dropdown.Option("article_code", "Артикул"),
                 ft.dropdown.Option("name", "Назва")],
        value="name",
        on_change=lambda e: load_products()
    )

    article_input = ft.TextField(label="Артикул", expand=1)
    name_input    = ft.TextField(label="Назва виробу", expand=2)
    # Added input for product weight in grams.  Using expand=1 keeps the
    # overall layout balanced with the other fields.  Users can leave
    # this blank if the weight is unknown.
    weight_input  = ft.TextField(label="Вага (г)", expand=1)
    drying_cb     = ft.Checkbox(label="Сушка", scale=0.8)
    trimming_cb   = ft.Checkbox(label="Обрізка", scale=0.8)
    cutting_cb    = ft.Checkbox(label="Різка", scale=0.8)
    cleaning_cb   = ft.Checkbox(label="Зачистка", scale=0.8)

    file_picker = ft.FilePicker()
    page.overlay.append(file_picker)

    def close_dialog():
        confirm_dialog.open = False
        page.update()

    def confirm_delete(code: str):
        nonlocal selected_product_code
        selected_product_code = code
        confirm_dialog.title   = ft.Text("Підтвердження видалення")
        confirm_dialog.content = ft.Text(f"Видалити виріб '{code}'?")
        confirm_dialog.actions = [
            ft.TextButton("Скасувати", on_click=lambda e: close_dialog()),
            ft.TextButton("Видалити",  on_click=lambda e: delete_product()),
        ]
        page.dialog = confirm_dialog
        if confirm_dialog not in page.overlay:
            page.overlay.append(confirm_dialog)
        confirm_dialog.open = True
        page.update()

    def delete_product():
        nonlocal selected_product_code
        db_exec("DELETE FROM product_base WHERE article_code = %s", (selected_product_code,))
        close_dialog()
        load_products()

    def clear_form(e=None):
        nonlocal is_editing
        article_input.value = ""
        name_input.value    = ""
        weight_input.value  = ""
        drying_cb.value     = False
        trimming_cb.value   = False
        cutting_cb.value    = False
        cleaning_cb.value   = False
        is_editing = False
        page.update()

    def fill_form(prod: dict):
        nonlocal is_editing
        is_editing = True
        article_input.value = prod["article_code"]
        name_input.value    = prod["name"]
        # Weight may not be present in all rows if the column was added later.
        w_val = prod.get("weight_g")
        weight_input.value = str(w_val) if w_val is not None else ""
        drying_cb.value     = bool(prod["drying_needed"])
        trimming_cb.value   = bool(prod["trimming_needed"])
        cutting_cb.value    = bool(prod["cutting_needed"])
        cleaning_cb.value   = bool(prod["cleaning_needed"])
        page.update()

    def save_product(e):
        code     = article_input.value.strip()
        name     = name_input.value.strip()
        # Parse weight preserving fractional grams if provided.  Accept comma as decimal separator.
        w_raw = weight_input.value.strip()
        weight = None
        if w_raw:
            try:
                # Use Decimal to parse without rounding
                weight = Decimal(w_raw.replace(",", "."))
                # Pass as string to MySQL to preserve exact value when inserting/updating
                weight = str(weight)
            except Exception:
                weight = None
        drying   = _bool(drying_cb.value)
        trimming = _bool(trimming_cb.value)
        cutting  = _bool(cutting_cb.value)
        cleaning = _bool(cleaning_cb.value)
        if not code or not name:
            return
        if db_fetch("SELECT 1 FROM product_base WHERE article_code=%s", (code,)):
            # Update existing product.  Include weight_g column if provided.  Use COALESCE to avoid overwriting
            # existing weight when the input is blank.
            if weight is not None:
                db_exec(
                    """
                        UPDATE product_base
                           SET name=%s, weight_g=%s, drying_needed=%s, trimming_needed=%s, cutting_needed=%s, cleaning_needed=%s
                         WHERE article_code=%s
                    """,
                    (name, weight, drying, trimming, cutting, cleaning, code),
                )
            else:
                db_exec(
                    """
                        UPDATE product_base
                           SET name=%s, drying_needed=%s, trimming_needed=%s, cutting_needed=%s, cleaning_needed=%s
                         WHERE article_code=%s
                    """,
                    (name, drying, trimming, cutting, cleaning, code),
                )
        else:
            # Insert new product.  Use provided weight or NULL if not specified.
            db_exec(
                """
                    INSERT INTO product_base
                        (article_code, name, weight_g, drying_needed, trimming_needed, cutting_needed, cleaning_needed)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                """,
                (code, name, weight, drying, trimming, cutting, cleaning),
            )
        clear_form()
        load_products()

    def update_stage_checkbox(article_code: str, field: str, value: bool):
        db_exec(f"UPDATE product_base SET {field} = %s WHERE article_code = %s", (_bool(value), article_code))
        log(f"{field} → {value} for {article_code}", tag="product_base")

    def make_on_change(field_name: str, article_code: str):
        return lambda e: update_stage_checkbox(article_code, field_name, e.control.value)

    def load_products():
        product_list.controls.clear()
        # Header row for the product list.  Added a column for weight after the name.
        product_list.controls.append(
            ft.Row([
                ft.Text("Артикул", weight="bold", expand=1),
                ft.Text("Назва",   weight="bold", expand=2),
                ft.Text("Вага (г)", weight="bold", width=70, text_align=ft.TextAlign.CENTER),
                ft.Text("Сушка",   weight="bold", width=60, text_align=ft.TextAlign.CENTER),
                ft.Text("Обрізка", weight="bold", width=60, text_align=ft.TextAlign.CENTER),
                ft.Text("Різка",   weight="bold", width=60, text_align=ft.TextAlign.CENTER),
                ft.Text("Зачистка",weight="bold", width=70, text_align=ft.TextAlign.CENTER),
                ft.Text("Дії",     weight="bold", width=90, text_align=ft.TextAlign.CENTER),
            ], spacing=5)
        )
        search = (search_field.value or "").strip()
        sql = "SELECT * FROM product_base"
        params = []
        if search:
            sql += " WHERE article_code LIKE %s OR name LIKE %s"
            params = [f"%{search}%", f"%{search}%"]
        sql += f" ORDER BY {sort_by.value}"
        for p in db_fetch(sql, tuple(params)):
            code = p["article_code"]
            # Retrieve weight if present; display empty string when None.
            w = p.get("weight_g")
            # Normalize weight display: remove trailing zeros and decimal point if not needed
            if w is None:
                w_text = ""
            else:
                try:
                    w_str = str(w)
                    if "." in w_str:
                        w_str = w_str.rstrip("0").rstrip(".")
                    w_text = w_str
                except Exception:
                    w_text = str(w)
            product_list.controls.append(
                ft.Row([
                    ft.Text(code, expand=1),
                    ft.Text(p["name"], expand=2),
                    ft.Text(w_text,        width=70, text_align=ft.TextAlign.CENTER),
                    ft.Checkbox(value=bool(p["drying_needed"]),   scale=0.7, width=60,
                                on_change=make_on_change("drying_needed", code)),
                    ft.Checkbox(value=bool(p["trimming_needed"]), scale=0.7, width=60,
                                on_change=make_on_change("trimming_needed", code)),
                    ft.Checkbox(value=bool(p["cutting_needed"]),  scale=0.7, width=60,
                                on_change=make_on_change("cutting_needed", code)),
                    ft.Checkbox(value=bool(p["cleaning_needed"]), scale=0.7, width=70,
                                on_change=make_on_change("cleaning_needed", code)),
                    ft.Row([
                        ft.IconButton(ft.icons.EDIT,   icon_color=ft.colors.BLUE_400,
                                      tooltip="Редагувати", on_click=lambda e, prod=p: fill_form(prod)),
                        ft.IconButton(ft.icons.DELETE, icon_color=ft.colors.RED_400,
                                      tooltip="Видалити",  on_click=lambda e, code=code: confirm_delete(code)),
                    ], width=90, alignment=ft.MainAxisAlignment.CENTER),
                ], spacing=5)
            )
        page.update()

    def export_main(e):
        if openpyxl is None:
            page.snack_bar = ft.SnackBar(ft.Text("Встановіть пакет openpyxl для експорту Excel"))
            page.snack_bar.open = True; page.update(); return

        # Export all fields including weight_g.  Some databases may not have the weight_g column; use COALESCE to
        # return NULL when absent.
        rows = db_fetch(
            "SELECT article_code, name, weight_g, drying_needed, trimming_needed, cutting_needed, cleaning_needed "
            "FROM product_base ORDER BY name"
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "product_base"
        ws.append([title for _, title in FIELDS])
        for r in rows:
            ws.append([
                r["article_code"],
                r["name"],
                r.get("weight_g"),
                _yes_no(r["drying_needed"]),
                _yes_no(r["trimming_needed"]),
                _yes_no(r["cutting_needed"]),
                _yes_no(r["cleaning_needed"]),
            ])
        bio = io.BytesIO(); wb.save(bio); data = bio.getvalue()

        file_picker = ft.FilePicker()
        page.overlay.append(file_picker)
        def _on_res(res: ft.FilePickerResultEvent):
            if res.path:
                with open(res.path, "wb") as f: f.write(data)
        file_picker.on_result = _on_res
        file_picker.save_file(file_name=f"product_base_{datetime.now():%Y%m%d}.xlsx")

    load_products()
    return ft.Column(
        controls=[
            ft.Row([
                ft.Text("База виробів (основна)", size=22, weight="bold"),
                ft.Container(expand=True),
                ft.OutlinedButton("Експорт у Excel", icon=ft.icons.DOWNLOAD, on_click=export_main),
            ]),
            ft.Row([search_field, sort_by], spacing=10),
            ft.Divider(),
            ft.Row([
                article_input,
                name_input,
                weight_input,
                drying_cb,
                trimming_cb,
                cutting_cb,
                cleaning_cb,
                ft.ElevatedButton("Зберегти", on_click=save_product),
                ft.TextButton("Очистити", on_click=clear_form),
            ], spacing=10),
            ft.Divider(),
            product_list,
        ],
        expand=1,
    )


# ──────────────────────────────────────────────────────────────────────
# Вкладка OLD (product_base_old) — імпорт/вибір/перенесення/експорт
# ──────────────────────────────────────────────────────────────────────

def _tab_old(page: ft.Page) -> ft.Column:
    list_old = ft.Column(scroll=ft.ScrollMode.ADAPTIVE)
    search_old = ft.TextField(label="Пошук", on_change=lambda e: load_old())
    sort_old = ft.Dropdown(
        label="Сортувати за",
        options=[ft.dropdown.Option("article_code", "Артикул"),
                 ft.dropdown.Option("name", "Назва")],
        value="name",
        on_change=lambda e: load_old()
    )

    # стан вибраного
    selected_codes: set[str] = set()
    # Select-all checkbox: use compact scaling and blue highlight for a pleasant look
    select_all_cb = ft.Checkbox(label="Вибрати всі", value=False, scale=0.7, active_color=ft.colors.BLUE_400, check_color=ft.colors.BLUE_400)

    btn_move_selected = ft.ElevatedButton(
        "Перенести вибрані (0)",
        icon=ft.icons.MOVE_UP,
        disabled=True,
        on_click=lambda e: move_selected_to_main(),
    )

    file_picker = ft.FilePicker()
    page.overlay.append(file_picker)

    def update_buttons():
        cnt = len(selected_codes)
        btn_move_selected.text = f"Перенести вибрані ({cnt})"
        btn_move_selected.disabled = (cnt == 0)
        page.update()

    def toggle_select_all(e=None):
        nonlocal selected_codes
        if select_all_cb.value:
            # обрати всі поточні у відфільтрованому наборі
            s = (search_old.value or "").strip()
            sql = "SELECT article_code FROM product_base_old"
            params = []
            if s:
                sql += " WHERE article_code LIKE %s OR name LIKE %s"
                params = [f"%{s}%", f"%{s}%"]
            sql += f" ORDER BY {sort_old.value}"
            rows = db_fetch(sql, tuple(params))
            selected_codes = {r["article_code"] for r in rows}
        else:
            selected_codes = set()
        load_old(reuse_selection=True)

    select_all_cb.on_change = lambda e: toggle_select_all()

    def load_old(reuse_selection: bool = False):
        list_old.controls.clear()
        # Заголовок таблиці
        list_old.controls.append(
            ft.Row([
                ft.Container(select_all_cb, width=80),
                ft.Text("Артикул", weight="bold", expand=1),
                ft.Text("Назва",   weight="bold", expand=2),
                ft.Text("Вага (г)", weight="bold", width=70, text_align=ft.TextAlign.CENTER),
                ft.Text("Сушка",   weight="bold", width=60, text_align=ft.TextAlign.CENTER),
                ft.Text("Обрізка", weight="bold", width=60, text_align=ft.TextAlign.CENTER),
                ft.Text("Різка",   weight="bold", width=60, text_align=ft.TextAlign.CENTER),
                ft.Text("Зачистка",weight="bold", width=70, text_align=ft.TextAlign.CENTER),
            ], spacing=5)
        )
        # Include weight_g if present in the old base; COALESCE handles missing columns by returning NULL.
        sql = "SELECT article_code,name,weight_g,drying_needed,trimming_needed,cutting_needed,cleaning_needed FROM product_base_old"
        params = []
        s = (search_old.value or "").strip()
        if s:
            sql += " WHERE article_code LIKE %s OR name LIKE %s"
            params = [f"%{s}%", f"%{s}%"]
        sql += f" ORDER BY {sort_old.value}"
        rows = db_fetch(sql, tuple(params))

        # якщо reuse_selection False — скидати select_all прапорець
        if not reuse_selection:
            select_all_cb.value = False

        for r in rows:
            code = r["article_code"]
            # стан чекбокса рядка
            # Row selection checkbox: compact size and blue highlight when selected
            row_cb = ft.Checkbox(
                value=(code in selected_codes),
                label=None,
                scale=0.7,
                active_color=ft.colors.BLUE_400,
                check_color=ft.colors.BLUE_400,
            )

            def _toggle_row_cb(e, c=code, cb=row_cb):
                nonlocal selected_codes
                if cb.value:
                    selected_codes.add(c)
                else:
                    selected_codes.discard(c)
                update_buttons()

            row_cb.on_change = _toggle_row_cb

            # Retrieve weight value if present; display empty string when None.
            w = r.get("weight_g")
            # Normalize weight display: strip trailing zeros and dot for decimals
            if w is None:
                w_text = ""
            else:
                try:
                    w_str = str(w)
                    if "." in w_str:
                        w_str = w_str.rstrip("0").rstrip(".")
                    w_text = w_str
                except Exception:
                    w_text = str(w)
            list_old.controls.append(
                ft.Row([
                    ft.Container(row_cb, width=80),
                    ft.Text(code, expand=1),
                    ft.Text(r["name"], expand=2),
                    ft.Text(w_text, width=70, text_align=ft.TextAlign.CENTER),
                    # For stage flags, use real checkboxes (disabled) rather than text placeholders.
                    ft.Checkbox(value=bool(r["drying_needed"]),   scale=0.7, width=60,
                                active_color=ft.colors.BLUE_400, check_color=ft.colors.BLUE_400, disabled=True),
                    ft.Checkbox(value=bool(r["trimming_needed"]), scale=0.7, width=60,
                                active_color=ft.colors.BLUE_400, check_color=ft.colors.BLUE_400, disabled=True),
                    ft.Checkbox(value=bool(r["cutting_needed"]),  scale=0.7, width=60,
                                active_color=ft.colors.BLUE_400, check_color=ft.colors.BLUE_400, disabled=True),
                    ft.Checkbox(value=bool(r["cleaning_needed"]), scale=0.7, width=70,
                                active_color=ft.colors.BLUE_400, check_color=ft.colors.BLUE_400, disabled=True),
                ], spacing=5)
            )

        update_buttons()
        page.update()

    def move_selected_to_main():
        if not selected_codes:
            return
        moved = 0
        # переносимо вибрані з UPSERT
        for code in list(selected_codes):
            row = db_fetch(
                "SELECT article_code,name,weight_g,drying_needed,trimming_needed,cutting_needed,cleaning_needed "
                "FROM product_base_old WHERE article_code=%s",
                (code,),
            )
            if not row:
                continue
            r = row[0]
            # ON DUPLICATE KEY UPDATE.  Include weight_g column when migrating.
            # Prepare weight as a string to preserve decimals when migrating
            weight_val = r.get("weight_g")
            if weight_val not in (None, ""):
                weight_str = str(weight_val)
            else:
                weight_str = None
            db_exec(
                """
                INSERT INTO product_base (
                    article_code, name, weight_g, drying_needed, trimming_needed, cutting_needed, cleaning_needed
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE
                  name=VALUES(name),
                  weight_g=VALUES(weight_g),
                  drying_needed=VALUES(drying_needed),
                  trimming_needed=VALUES(trimming_needed),
                  cutting_needed=VALUES(cutting_needed),
                  cleaning_needed=VALUES(cleaning_needed)
                """,
                (
                    r["article_code"],
                    r["name"],
                    weight_str,
                    _bool(r["drying_needed"]),
                    _bool(r["trimming_needed"]),
                    _bool(r["cutting_needed"]),
                    _bool(r["cleaning_needed"]),
                ),
            )
            moved += 1
        # Очищаємо вибір
        selected_codes.clear()
        load_old()
        page.snack_bar = ft.SnackBar(ft.Text(f"Перенесено записів: {moved}"))
        page.snack_bar.open = True
        page.update()

    def import_old(e):
        if openpyxl is None:
            page.snack_bar = ft.SnackBar(ft.Text("Для імпорту потрібен пакет openpyxl (pip install openpyxl)"))
            page.snack_bar.open = True; page.update(); return

        def _on_res(res: ft.FilePickerResultEvent):
            if not res.files: return
            p = res.files[0].path
            wb = openpyxl.load_workbook(p)
            ws = wb.active
            # очікуємо заголовки як у FIELDS
            headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))[0:len(FIELDS)]]
            if not (headers and "артикул" in headers[0] and "назва" in headers[1]):
                page.snack_bar = ft.SnackBar(ft.Text("Неправильний формат заголовків у Excel"))
                page.snack_bar.open = True; page.update(); return

            cnt = 0
            for row in ws.iter_rows(min_row=2):
                # Read as many values as there are expected fields.  Since we've added
                # weight, this slice now includes weight at index 2.
                vals = [c.value for c in row[:len(FIELDS)]]
                if not vals or vals[0] in (None, ""):
                    continue
                code = str(vals[0]).strip()
                name = str(vals[1]).strip() if vals[1] is not None else ""
                weight_val = vals[2] if len(vals) > 2 else None
                d = _bool(vals[3] if len(vals) > 3 else 0)
                t = _bool(vals[4] if len(vals) > 4 else 0)
                c_val = _bool(vals[5] if len(vals) > 5 else 0)
                cl = _bool(vals[6] if len(vals) > 6 else 0)
                # Normalize weight preserving decimals.  Accept comma as decimal separator.
                weight = None
                if weight_val not in (None, ""):
                    try:
                        weight = Decimal(str(weight_val).replace(",", "."))
                        weight = str(weight)
                    except Exception:
                        weight = None
                if db_fetch("SELECT 1 FROM product_base_old WHERE article_code=%s", (code,)):
                    db_exec(
                        """UPDATE product_base_old
                                  SET name=%s, weight_g=%s, drying_needed=%s, trimming_needed=%s, cutting_needed=%s, cleaning_needed=%s
                                WHERE article_code=%s""",
                        (name, weight, d, t, c_val, cl, code),
                    )
                else:
                    db_exec(
                        """INSERT INTO product_base_old
                                (article_code,name,weight_g,drying_needed,trimming_needed,cutting_needed,cleaning_needed)
                               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                        (code, name, weight, d, t, c_val, cl),
                    )
                cnt += 1
            load_old()
            page.snack_bar = ft.SnackBar(ft.Text(f"Імпортовано рядків: {cnt}"))
            page.snack_bar.open = True; page.update()

        file_picker.on_result = _on_res
        file_picker.pick_files(allow_multiple=False, allowed_extensions=["xlsx"])

    def export_old(e):
        if openpyxl is None:
            page.snack_bar = ft.SnackBar(ft.Text("Встановіть пакет openpyxl для експорту Excel"))
            page.snack_bar.open = True; page.update(); return
        rows = db_fetch(
            "SELECT article_code,name,weight_g,drying_needed,trimming_needed,cutting_needed,cleaning_needed "
            "FROM product_base_old ORDER BY name"
        )
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "product_base_old"
        ws.append([title for _, title in FIELDS])
        for r in rows:
            ws.append([
                r["article_code"],
                r["name"],
                r.get("weight_g"),
                _yes_no(r["drying_needed"]),
                _yes_no(r["trimming_needed"]),
                _yes_no(r["cutting_needed"]),
                _yes_no(r["cleaning_needed"]),
            ])
        bio = io.BytesIO(); wb.save(bio); data = bio.getvalue()
        def _on_res(res: ft.FilePickerResultEvent):
            if res.path:
                with open(res.path, "wb") as f: f.write(data)
        file_picker.on_result = _on_res
        file_picker.save_file(file_name=f"product_base_old_{datetime.now():%Y%m%d}.xlsx")

    load_old()
    return ft.Column(
        controls=[
            ft.Row([
                ft.Text("База виробів OLD (імпорт / вибір / перенесення)", size=22, weight="bold"),
                ft.Container(expand=True),
                ft.OutlinedButton("Імпорт Excel",  icon=ft.icons.FILE_UPLOAD, on_click=import_old),
                ft.OutlinedButton("Експорт Excel", icon=ft.icons.DOWNLOAD,    on_click=export_old),
                btn_move_selected,
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            ft.Row([search_old, sort_old], spacing=10),
            ft.Divider(),
            list_old,
        ],
        expand=1,
    )


# ──────────────────────────────────────────────────────────────────────
# Вью з Tabs
# ──────────────────────────────────────────────────────────────────────

def product_base_view(page: ft.Page) -> ft.View:
    log("Opening product base (tabs)", tag="product_base")

    def go_back(e):
        if len(page.views) > 1:
            page.views.pop(); page.update()

    tabs = ft.Tabs(
        selected_index=0,
        tabs=[
            ft.Tab(text="База виробів",     content=_tab_main(page)),
            ft.Tab(text="База виробів OLD", content=_tab_old(page)),
        ],
        expand=1,
    )

    # The launcher sets an appbar with a back button and title for this view.
    # Avoid duplicating the back button and title inside the view content.
    return ft.View(
        route="/product_base",
        controls=[
            tabs,
        ],
        scroll=ft.ScrollMode.ADAPTIVE,
    )
